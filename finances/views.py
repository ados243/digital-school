from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps
from io import BytesIO

from django.db import transaction
from django.db.models import Sum, Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify


def _is_ajax(request):
    return (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in (request.headers.get("Accept") or "")
    )


def _type_frais_payload(obj):
    return {
        "id": obj.id,
        "libelle": obj.libelle,
        "description": obj.description or "",
        "label": obj.libelle,
    }


def _peut_encaisser(user):
    """Seul un caissier (rôle ou fonction GRH) peut exécuter un paiement."""
    return bool(
        user
        and user.is_authenticated
        and getattr(user, "is_caissier", False)
    )


def _peut_gerer_ecritures(user):
    """Seuls trésorier / comptable / superuser peuvent passer ou modifier les écritures."""
    return bool(
        user
        and user.is_authenticated
        and getattr(user, "peut_gerer_ecritures", False)
    )


def _peut_modifier_paiement(user):
    """Les caissiers ne modifient ni ne suppriment les paiements."""
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    if getattr(user, "is_caissier", False):
        return False
    return bool(getattr(user, "is_personnel_interne", False))


def _demande_modification_ouverte(paiement):
    return DemandeModificationPaiement.ouverte_pour_paiement(paiement)


def _peut_modifier_ce_paiement(user, paiement):
    """
    L'admin (personnel interne) ne peut modifier que s'il existe une demande EN_ATTENTE.
    Superuser : même règle métier (sauf si on veut un bypass — on applique la règle).
    """
    if not _peut_modifier_paiement(user):
        return False
    return _demande_modification_ouverte(paiement) is not None


def _nom_acteur(user):
    return (
        getattr(user, "prenom", "")
        or user.get_full_name()
        or user.username
    )


def _identifiants_caissier(user):
    """Noms possibles enregistrés dans Paiement.caissier pour l'utilisateur."""
    noms = set()
    for v in (
        _nom_acteur(user),
        getattr(user, "prenom", "") or "",
        user.get_full_name() if hasattr(user, "get_full_name") else "",
        getattr(user, "username", "") or "",
    ):
        v = (v or "").strip()
        if v:
            noms.add(v)
    return noms


def caissier_required(view_func):
    """Restreint l'exécution (encaissement) des paiements au seul caissier."""

    @login_required
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not _peut_encaisser(request.user):
            messages.error(
                request,
                "Seul un caissier peut effectuer un paiement.",
            )
            return redirect("finances:paiement_list")
        return view_func(request, *args, **kwargs)

    return _wrapped


def modification_paiement_required(view_func):
    """Interdit modification / suppression des paiements aux caissiers."""

    @login_required
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not _peut_modifier_paiement(request.user):
            messages.error(
                request,
                "Vous n'êtes pas autorisé à modifier ou supprimer un paiement.",
            )
            return redirect("finances:paiement_list")
        return view_func(request, *args, **kwargs)

    return _wrapped


from .forms import (
    TypeFraisForm,
    FraisScolaireForm,
    PaiementForm,
    TauxChangeForm,
    DemandeModificationPaiementForm,
    ConfigWhatsAppForm,
    # HOADA
    CompteComptableForm,
    JournalComptableForm,
    EcritureForm,
    EcritureLigneForm,
)
from .models import (
    Frais_Scolaire,
    Paiement,
    TypeFrais,
    Devise,
    TauxChange,
    DemandeModificationPaiement,
    ConfigWhatsApp,
    NotificationWhatsApp,
    ClotureCaisse,
    BudgetAnnuel,
    LigneBudget,
    PosteBudget,
    # HOADA
    CompteComptable,
    JournalComptable,
    PieceComptable,
    Ecriture,
    EcritureLigne,
)
from .tenant import get_user_ecole, type_frais_for_ecole, frais_for_ecole, paiements_for_ecole, taux_change_for_ecole
from inscription.models import Classe
from inscription.tenant import (
    inscriptions_for_ecole,
    annees_for_ecole,
    classes_for_ecole,
    sections_for_ecole,
)
from .paiement_utils import (
    _decimal,
    frais_disponibles_pour_inscription,
    build_frais_solde_context,
    solde_caisse_comptable,
    caisse_especes_par_devise,
    types_frais_minerval_for_ecole,
    minerval_par_classe,
    minerval_paiements_queryset,
    projection_budget_minerval,
    construire_postes_budget,
)


def _parse_optional_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _parse_optional_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _situation_encaissements(paiements_qs, date_debut, date_fin):
    """Agrège les paiements VALIDÉS sur une période (dates inclusives)."""
    qs = paiements_qs.filter(
        statut="VALIDE",
        date_encodage__date__gte=date_debut,
        date_encodage__date__lte=date_fin,
    )
    par_devise = list(
        qs.values("devise__devise")
        .annotate(total=Sum("montant_paye"))
        .order_by("-total")
    )
    return {
        "nb": qs.count(),
        "par_devise": par_devise,
        "date_debut": date_debut,
        "date_fin": date_fin,
    }


def _situation_depenses(paies_qs, date_debut, date_fin):
    """Agrège les salaires PAYÉS (dépenses) sur une période (dates inclusives)."""
    qs = paies_qs.filter(
        statut_paiement="PAYE",
        date_paiement__gte=date_debut,
        date_paiement__lte=date_fin,
    )
    par_devise = list(
        qs.values("devise__devise")
        .annotate(total=Sum("net_a_payer"))
        .order_by("-total")
    )
    return {
        "nb": qs.count(),
        "par_devise": par_devise,
        "date_debut": date_debut,
        "date_fin": date_fin,
    }


@login_required
def dashboard(request):
    ecole = get_user_ecole(request)
    paiements_qs = paiements_for_ecole(ecole)
    frais_qs = frais_for_ecole(ecole)

    total_encaisse_par_devise = (
        paiements_qs.filter(statut="VALIDE")
        .values("devise__devise")
        .annotate(total=Sum("montant_paye"))
        .order_by("-total")
    )

    # Situations journalière / hebdomadaire / mensuelle (entrées + dépenses)
    from grh.tenant import paies_for_ecole

    today = timezone.localdate()
    debut_semaine = today - timedelta(days=today.weekday())  # lundi
    debut_mois = today.replace(day=1)
    situation_jour = _situation_encaissements(paiements_qs, today, today)
    situation_semaine = _situation_encaissements(paiements_qs, debut_semaine, today)
    situation_mois = _situation_encaissements(paiements_qs, debut_mois, today)

    paies_qs = paies_for_ecole(ecole)
    depenses_jour = _situation_depenses(paies_qs, today, today)
    depenses_semaine = _situation_depenses(paies_qs, debut_semaine, today)
    depenses_mois = _situation_depenses(paies_qs, debut_mois, today)

    frais_par_type = (
        frais_qs.values("type_frais__libelle")
        .annotate(nb=Count("id"), montant_total=Sum("montant"))
        .order_by("-nb")
    )

    recent_paiements = (
        paiements_qs.select_related(
            "eleve", "frais", "devise", "devise_origine", "eleve__eleve"
        )
        .order_by("-date_encodage")[:8]
    )

    recent_frais = (
        frais_qs.select_related("type_frais", "annee", "section", "devise")
        .order_by("-echeance")[:8]
    )

    # --- Caisse réelle ---
    solde_caisse = solde_caisse_comptable(ecole)
    caisse_par_devise = caisse_especes_par_devise(ecole)

    # --- Filtres minerval ---
    filter_annee = _parse_optional_int(request.GET.get("annee"))
    filter_section = _parse_optional_int(request.GET.get("section"))
    filter_classe = _parse_optional_int(request.GET.get("classe"))
    filter_devise = _parse_optional_int(request.GET.get("devise"))
    filter_type_frais = _parse_optional_int(request.GET.get("type_frais"))
    filter_mode = (request.GET.get("mode_paiement") or "").strip() or None
    filter_date_debut = _parse_optional_date(request.GET.get("date_debut"))
    filter_date_fin = _parse_optional_date(request.GET.get("date_fin"))

    minerval_filters = {
        "annee": filter_annee,
        "section": filter_section,
        "classe": filter_classe,
        "devise": filter_devise,
        "type_frais": filter_type_frais,
        "mode_paiement": filter_mode,
        "date_debut": filter_date_debut,
        "date_fin": filter_date_fin,
    }

    minerval_rows = minerval_par_classe(ecole, minerval_filters)
    minerval_total = (
        minerval_paiements_queryset(ecole, minerval_filters)
        .aggregate(total=Sum("montant_paye"))["total"]
        or Decimal("0")
    )

    annees = annees_for_ecole(ecole).order_by("-est_encoure", "-anne_scolaire")
    sections = sections_for_ecole(ecole).order_by("section")
    classes_qs = classes_for_ecole(ecole).select_related("section").order_by("section__section", "classe")
    if filter_section:
        classes_qs = classes_qs.filter(section_id=filter_section)

    # Querystring des filtres (pour liens vers détail classe)
    filter_query = request.GET.urlencode()

    context = {
        "total_encaisse_par_devise": total_encaisse_par_devise,
        "frais_par_type": frais_par_type,
        "recent_paiements": recent_paiements,
        "recent_frais": recent_frais,
        # Situations périodiques (entrées)
        "situation_jour": situation_jour,
        "situation_semaine": situation_semaine,
        "situation_mois": situation_mois,
        # Situations périodiques (dépenses / salaires payés)
        "depenses_jour": depenses_jour,
        "depenses_semaine": depenses_semaine,
        "depenses_mois": depenses_mois,
        # Caisse
        "solde_caisse": solde_caisse,
        "caisse_par_devise": caisse_par_devise,
        # Minerval
        "minerval_rows": minerval_rows,
        "minerval_total": minerval_total,
        "filter_query": filter_query,
        # Filtres
        "filter_annee": filter_annee,
        "filter_section": filter_section,
        "filter_classe": filter_classe,
        "filter_devise": filter_devise,
        "filter_type_frais": filter_type_frais,
        "filter_mode": filter_mode or "",
        "filter_date_debut": filter_date_debut.isoformat() if filter_date_debut else "",
        "filter_date_fin": filter_date_fin.isoformat() if filter_date_fin else "",
        "annees": annees,
        "sections": sections,
        "classes": classes_qs,
        "devises": Devise.objects.all().order_by("devise"),
        "types_minerval": types_frais_minerval_for_ecole(ecole).order_by("libelle"),
        "modes_paiement": Paiement.mode_paiement_choices,
    }

    return render(request, "finances/dashboard.html", context)


def _dashboard_filters_from_request(request):
    return {
        "annee": _parse_optional_int(request.GET.get("annee")),
        "section": _parse_optional_int(request.GET.get("section")),
        "classe": _parse_optional_int(request.GET.get("classe")),
        "devise": _parse_optional_int(request.GET.get("devise")),
        "type_frais": _parse_optional_int(request.GET.get("type_frais")),
        "mode_paiement": (request.GET.get("mode_paiement") or "").strip() or None,
        "date_debut": _parse_optional_date(request.GET.get("date_debut")),
        "date_fin": _parse_optional_date(request.GET.get("date_fin")),
        "statut": (request.GET.get("statut") or "").strip() or None,
    }


def _paiements_for_classe(ecole, classe, filters=None):
    """Paiements des élèves inscrits dans une classe, avec filtres optionnels."""
    filters = filters or {}
    inscriptions = (
        inscriptions_for_ecole(ecole)
        .filter(classe=classe)
        .select_related("eleve", "classe", "annee_s")
    )
    if filters.get("annee"):
        inscriptions = inscriptions.filter(annee_s_id=filters["annee"])

    qs = (
        paiements_for_ecole(ecole)
        .filter(eleve__in=inscriptions)
        .select_related(
            "eleve",
            "eleve__eleve",
            "eleve__classe",
            "eleve__annee_s",
            "frais",
            "frais__type_frais",
            "devise",
        )
        .order_by("eleve__eleve__nom", "eleve__eleve__prenom", "-date_encodage")
    )

    if filters.get("devise"):
        qs = qs.filter(devise_id=filters["devise"])
    if filters.get("type_frais"):
        qs = qs.filter(frais__type_frais_id=filters["type_frais"])
    if filters.get("mode_paiement"):
        qs = qs.filter(mode_paiement=filters["mode_paiement"])
    if filters.get("date_debut"):
        qs = qs.filter(date_encodage__date__gte=filters["date_debut"])
    if filters.get("date_fin"):
        qs = qs.filter(date_encodage__date__lte=filters["date_fin"])
    if filters.get("statut"):
        qs = qs.filter(statut=filters["statut"])

    return inscriptions, qs


def _export_classe_paiements_excel(classe, inscriptions, paiements, totaux_par_devise):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from .paiement_utils import paiements_valides_par_frais, solde_frais

    ecole = classe.ecole
    paye_par_frais = paiements_valides_par_frais(ecole)

    wb = Workbook()
    ws = wb.active
    ws.title = "Paiements"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )

    ws["A1"] = f"Liste des paiements — Classe {classe.classe}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Section : {classe.section}"
    ws["A3"] = f"Élèves inscrits : {inscriptions.count()}"
    ws["A4"] = f"Exporté le : {timezone.now().strftime('%d/%m/%Y %H:%M')}"

    headers = [
        "Nom",
        "Post-nom",
        "Prénom",
        "Type de frais",
        "Montant",
        "Devise",
        "Date",
        "Montant restant",
        "Solde",
    ]
    start_row = 6
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for offset, p in enumerate(paiements):
        eleve = p.eleve.eleve
        solde = solde_frais(p.frais, p.eleve_id, paye_par_frais) if p.frais_id else {
            "reste": Decimal("0"),
            "est_solde": True,
        }
        montant_restant = _decimal(solde["reste"])
        row = start_row + 1 + offset
        values = [
            eleve.nom,
            eleve.Post_nom or "",
            eleve.prenom,
            p.frais.type_frais.libelle if p.frais_id else "",
            float(p.montant_paye or 0),
            str(p.devise) if p.devise_id else "",
            p.date_encodage.strftime("%d/%m/%Y") if p.date_encodage else "",
            float(montant_restant),
            "Soldé" if solde["est_solde"] else "Non soldé",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin

    totals_row = start_row + 1 + len(paiements) + 1
    ws.cell(row=totals_row, column=1, value="Totaux par devise").font = Font(bold=True)
    for i, row in enumerate(totaux_par_devise):
        ws.cell(row=totals_row + 1 + i, column=1, value=row["devise__devise"])
        ws.cell(row=totals_row + 1 + i, column=2, value=float(row["total"] or 0))

    # Feuille synthèse élèves
    ws2 = wb.create_sheet("Élèves")
    ws2["A1"] = "Synthèse par élève"
    ws2["A1"].font = Font(bold=True, size=13)
    eleve_headers = [
        "Nom",
        "Post-nom",
        "Prénom",
        "Nb paiements",
        "Total payé",
        "Devise",
        "Montant restant",
        "Solde",
    ]
    for col, title in enumerate(eleve_headers, start=1):
        cell = ws2.cell(row=3, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin

    resume = {}
    for p in paiements:
        key = (p.eleve_id, str(p.devise) if p.devise_id else "")
        if key not in resume:
            eleve = p.eleve.eleve
            resume[key] = {
                "inscription_id": p.eleve_id,
                "nom": eleve.nom,
                "post_nom": eleve.Post_nom or "",
                "prenom": eleve.prenom,
                "nb": 0,
                "total": Decimal("0"),
                "devise": str(p.devise) if p.devise_id else "",
                "frais_ids": set(),
            }
        resume[key]["nb"] += 1
        resume[key]["total"] += _decimal(p.montant_paye)
        if p.frais_id:
            resume[key]["frais_ids"].add(p.frais_id)

    payeurs_ids = {p.eleve_id for p in paiements}
    frais_by_id = {p.frais_id: p.frais for p in paiements if p.frais_id}
    row_idx = 4
    for item in sorted(resume.values(), key=lambda x: (x["nom"], x["prenom"])):
        montant_restant = Decimal("0")
        for frais_id in item["frais_ids"]:
            frais = frais_by_id.get(frais_id)
            if not frais:
                continue
            solde = solde_frais(frais, item["inscription_id"], paye_par_frais)
            montant_restant += max(solde["reste"], Decimal("0"))
        values = [
            item["nom"],
            item["post_nom"],
            item["prenom"],
            item["nb"],
            float(item["total"]),
            item["devise"],
            float(montant_restant),
            "Soldé" if montant_restant <= 0 else "Non soldé",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = thin
        row_idx += 1

    for ins in inscriptions.select_related("eleve").order_by("eleve__nom", "eleve__prenom"):
        if ins.id in payeurs_ids:
            continue
        eleve = ins.eleve
        values = [
            eleve.nom,
            eleve.Post_nom or "",
            eleve.prenom,
            0,
            0,
            "",
            0,
            "Non soldé",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = thin
        row_idx += 1

    for sheet in (ws, ws2):
        for column_cells in sheet.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"paiements_classe_{slugify(classe.classe) or classe.id}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def classe_paiements(request, classe_id):
    """Liste des paiements de tous les élèves inscrits dans une classe."""
    ecole = get_user_ecole(request)
    classe = get_object_or_404(
        Classe.objects.select_related("section", "ecole"),
        pk=classe_id,
        ecole=ecole,
    )
    filters = _dashboard_filters_from_request(request)
    # La classe est imposée par l'URL
    filters["classe"] = classe.id

    inscriptions, paiements = _paiements_for_classe(ecole, classe, filters)
    paiements_list = list(paiements)

    totaux_par_devise = list(
        paiements.filter(statut="VALIDE")
        .values("devise__devise")
        .annotate(total=Sum("montant_paye"))
        .order_by("-total")
    )
    total_general = sum((_decimal(r["total"]) for r in totaux_par_devise), Decimal("0"))

    if request.GET.get("export") == "excel":
        return _export_classe_paiements_excel(
            classe, inscriptions, paiements_list, totaux_par_devise
        )

    # Synthèse par élève + détails pour la popup
    resume_eleves = {}
    for p in paiements_list:
        key = p.eleve_id
        if key not in resume_eleves:
            resume_eleves[key] = {
                "nb": 0,
                "total": Decimal("0"),
                "paiements": [],
            }
        resume_eleves[key]["nb"] += 1
        if p.statut == "VALIDE":
            resume_eleves[key]["total"] += _decimal(p.montant_paye)
        resume_eleves[key]["paiements"].append({
            "id": p.id,
            "numero_recu": p.numero_recu or "",
            "frais": p.frais.type_frais.libelle if p.frais_id else "",
            "montant": str(p.montant_paye or 0),
            "devise": str(p.devise) if p.devise_id else "",
            "mode": p.get_mode_paiement_display(),
            "statut": p.statut,
            "statut_label": p.get_statut_display(),
            "date": p.date_encodage.strftime("%d/%m/%Y") if p.date_encodage else "",
            "print_url": (
                reverse("finances:paiement_print", args=[p.id])
                if p.statut == "VALIDE" and p.numero_recu
                else ""
            ),
        })

    eleves_rows = []
    paiements_by_eleve = {}
    for ins in inscriptions.select_related("eleve", "annee_s").order_by("eleve__nom", "eleve__prenom"):
        data = resume_eleves.get(ins.id)
        eleves_rows.append({
            "inscription": ins,
            "eleve": ins.eleve,
            "nb": data["nb"] if data else 0,
            "total": data["total"] if data else Decimal("0"),
        })
        paiements_by_eleve[str(ins.id)] = {
            "nom": f"{ins.eleve.nom} {ins.eleve.Post_nom} {ins.eleve.prenom}".strip(),
            "matricule": ins.eleve.matricule or "",
            "annee": str(ins.annee_s),
            "total": str(data["total"]) if data else "0",
            "paiements": data["paiements"] if data else [],
        }

    # Conserver les filtres dans les liens export/print
    query = request.GET.copy()
    query.pop("export", None)
    filter_query = query.urlencode()

    context = {
        "classe": classe,
        "inscriptions": inscriptions,
        "paiements": paiements_list,
        "eleves_rows": eleves_rows,
        "paiements_by_eleve": paiements_by_eleve,
        "totaux_par_devise": totaux_par_devise,
        "total_general": total_general,
        "nb_eleves": inscriptions.count(),
        "nb_paiements": len(paiements_list),
        "filter_query": filter_query,
        "filters": filters,
        "print_mode": request.GET.get("print") == "1",
    }
    return render(request, "finances/classe_paiements.html", context)


# -------------------------
# TypeFrais CRUD
# -------------------------

@login_required
def type_frais_list(request):
    ecole = get_user_ecole(request)
    items = type_frais_for_ecole(ecole).order_by("libelle")
    return render(request, "finances/type_frais_list.html", {"object_list": items})


@login_required
def type_frais_create(request):
    ecole = get_user_ecole(request)
    form_prefix = (
        "type_frais_modal"
        if request.method == "POST"
        and any(k.startswith("type_frais_modal-") for k in request.POST)
        else None
    )
    if request.method == "POST":
        form = TypeFraisForm(request.POST, prefix=form_prefix)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = ecole
            obj.save()
            if _is_ajax(request):
                return JsonResponse({"ok": True, "type_frais": _type_frais_payload(obj)})
            return redirect("finances:type_frais_list")
        if _is_ajax(request):
            errors = {
                field: [str(e) for e in errs]
                for field, errs in form.errors.items()
            }
            return JsonResponse({"ok": False, "errors": errors}, status=400)
    else:
        form = TypeFraisForm()
    return render(request, "finances/type_frais_form.html", {"form": form})


@login_required
def type_frais_update(request, pk):
    ecole = get_user_ecole(request)
    obj = get_object_or_404(TypeFrais, pk=pk, ecole=ecole)
    if request.method == "POST":
        form = TypeFraisForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("finances:type_frais_list")
    else:
        form = TypeFraisForm(instance=obj)
    return render(request, "finances/type_frais_update_form.html", {"form": form, "object": obj})


@login_required
def type_frais_delete(request, pk):
    ecole = get_user_ecole(request)
    obj = get_object_or_404(TypeFrais, pk=pk, ecole=ecole)
    if request.method == "POST":
        obj.delete()
        return redirect("finances:type_frais_list")
    return render(request, "finances/type_frais_confirm_delete.html", {"object": obj})


# -------------------------
# Frais_Scolaire CRUD
# -------------------------

@login_required
def frais_scolaire_list(request):
    ecole = get_user_ecole(request)
    items = frais_for_ecole(ecole).select_related("type_frais", "annee", "section", "devise").order_by("-echeance")
    return render(request, "finances/frais_scolaire_list.html", {"object_list": items})


@login_required
def frais_scolaire_create(request):
    ecole = get_user_ecole(request)
    if request.method == "POST":
        form = FraisScolaireForm(request.POST, ecole=ecole)
        if form.is_valid():
            form.save()
            return redirect("finances:frais_scolaire_list")
    else:
        form = FraisScolaireForm(ecole=ecole)

    from inscription.tenant import sections_for_ecole

    context = {
        "form": form,
        "type_frais_form": TypeFraisForm(prefix="type_frais_modal"),
        "type_frais_create_url": reverse("finances:type_frais_create"),
        "types_count": type_frais_for_ecole(ecole).count(),
        "sections_count": sections_for_ecole(ecole).count(),
        "annee_courante": annees_for_ecole(ecole).filter(est_encoure=True).first(),
    }
    return render(request, "finances/frais_scolaire_form.html", context)


@login_required
def frais_scolaire_update(request, pk):
    ecole = get_user_ecole(request)
    obj = get_object_or_404(frais_for_ecole(ecole), pk=pk)
    if request.method == "POST":
        form = FraisScolaireForm(request.POST, instance=obj, ecole=ecole)
        if form.is_valid():
            form.save()
            return redirect("finances:frais_scolaire_list")
    else:
        form = FraisScolaireForm(instance=obj, ecole=ecole)
    context = {
        "form": form,
        "object": obj,
        "type_frais_form": TypeFraisForm(prefix="type_frais_modal"),
        "type_frais_create_url": reverse("finances:type_frais_create"),
    }
    return render(request, "finances/frais_scolaire_update_form.html", context)


@login_required
def frais_scolaire_delete(request, pk):
    ecole = get_user_ecole(request)
    obj = get_object_or_404(frais_for_ecole(ecole), pk=pk)
    if request.method == "POST":
        obj.delete()
        return redirect("finances:frais_scolaire_list")
    return render(request, "finances/frais_scolaire_confirm_delete.html", {"object": obj})


# -------------------------
# Paiement CRUD
# -------------------------

def _paiement_form_context(request, ecole, form, is_update=False, paiement=None):
    inscriptions_qs = inscriptions_for_ecole(ecole).select_related("eleve", "classe", "annee_s")
    exclude_id = paiement.pk if paiement else None

    inscriptions_data = [
        {
            "id": ins.id,
            "nom": f"{ins.eleve.prenom} {ins.eleve.nom}",
            "matricule": ins.eleve.matricule,
            "classe": str(ins.classe),
            "annee": str(ins.annee_s),
            "classe_id": ins.classe_id,
            "annee_id": ins.annee_s_id,
            "label": f"{ins.eleve.prenom} {ins.eleve.nom} — {ins.classe} ({ins.annee_s}) · {ins.eleve.matricule}",
        }
        for ins in inscriptions_qs.order_by("eleve__nom", "eleve__prenom")
    ]
    frais_solde_by_inscription = build_frais_solde_context(ecole, inscriptions_qs, exclude_id)

    frais_count = sum(len(v) for v in frais_solde_by_inscription.values())
    taux_obj = TauxChange.courant_pour_ecole(ecole)

    return {
        "form": form,
        "is_update": is_update,
        "paiement": paiement,
        "inscriptions_data": inscriptions_data,
        "frais_solde_by_inscription": frais_solde_by_inscription,
        "inscriptions_count": inscriptions_qs.count(),
        "frais_count": frais_count,
        "annee_courante": annees_for_ecole(ecole).filter(est_encoure=True).first(),
        "taux_courant": taux_obj.taux if taux_obj else None,
        "taux_courant_obj": taux_obj,
    }


@login_required
def taux_change_list(request):
    """Historique et saisie du taux de conversion CDF ↔ USD."""
    ecole = get_user_ecole(request)
    qs = taux_change_for_ecole(ecole)
    courant = TauxChange.courant_pour_ecole(ecole)

    if request.method == "POST":
        form = TauxChangeForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = ecole
            obj.saisi_par = (
                getattr(request.user, "prenom", "")
                or request.user.get_full_name()
                or request.user.username
            )
            obj.save()
            messages.success(
                request,
                f"Taux enregistré : 1 USD = {obj.taux} CDF (effet {obj.date_effet:%d/%m/%Y}).",
            )
            return redirect("finances:taux_change_list")
    else:
        form = TauxChangeForm(
            initial={
                "date_effet": timezone.localdate(),
                "taux": courant.taux if courant else None,
            }
        )

    return render(
        request,
        "finances/taux_change_list.html",
        {
            "object_list": qs[:50],
            "form": form,
            "taux_courant": courant,
            "ecole": ecole,
        },
    )


@login_required
def whatsapp_config(request):
    """Configuration WhatsApp centrale + journal des envois."""
    ecole = get_user_ecole(request)
    if not request.user.is_superuser:
        messages.error(request, "Seul un superutilisateur peut configurer WhatsApp.")
        return redirect("finances:dashboard")

    config = ConfigWhatsApp.charger_centrale()

    if request.method == "POST" and request.POST.get("action") == "save_config":
        form = ConfigWhatsAppForm(request.POST, instance=config)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = None
            obj.save()
            messages.success(request, "Configuration WhatsApp centrale enregistrée.")
            return redirect("finances:whatsapp_config")
    else:
        form = ConfigWhatsAppForm(instance=config)

    from .whatsapp import CLES_CONTEXTE, parser_cles_template

    notif_qs = NotificationWhatsApp.objects.select_related("paiement", "ecole")
    if not request.user.is_superuser and ecole:
        notif_qs = notif_qs.filter(ecole=ecole)
    notifications = notif_qs.order_by("-date_envoi")[:80]
    mapping_vars = [
        (f"{{{{{i}}}}}", cle)
        for i, cle in enumerate(parser_cles_template(config), start=1)
    ]
    return render(
        request,
        "finances/whatsapp_config.html",
        {
            "form": form,
            "config": config,
            "notifications": notifications,
            "ecole": ecole,
            "journal_multi_ecoles": request.user.is_superuser,
            "placeholders": " ".join("{" + c + "}" for c in CLES_CONTEXTE),
            "mapping_vars": mapping_vars,
            "exemple_template_meta": (
                "Bonjour, paiement reçu pour {{1}}.\n"
                "Montant : {{2}}\n"
                "Reçu : {{3}}\n"
                "Frais : {{4}}\n"
                "Classe : {{5}}\n"
                "Date : {{6}}"
            ),
        },
    )


@login_required
def whatsapp_test(request):
    """Envoie un message de test au numéro saisi (ou au téléphone de l'école)."""
    ecole = get_user_ecole(request)
    if not request.user.is_superuser:
        messages.error(request, "Seul un superutilisateur peut tester WhatsApp.")
        return redirect("finances:dashboard")
    if request.method != "POST":
        return redirect("finances:whatsapp_config")

    config = ConfigWhatsApp.charger_centrale()
    if not config or not config.actif:
        messages.error(request, "Activez d'abord WhatsApp et enregistrez la configuration centrale.")
        return redirect("finances:whatsapp_config")

    from .whatsapp import (
        contexte_test,
        formater_message,
        normaliser_telephone,
        resume_envoi_meta,
        _envoyer_via_provider,
    )

    tel_brut = (request.POST.get("telephone_test") or "").strip() or (
        getattr(ecole, "telephone1", None) or ""
    )
    telephone = normaliser_telephone(tel_brut, config.indicatif_pays)
    if not telephone:
        messages.error(request, "Numéro de test invalide.")
        return redirect("finances:whatsapp_config")

    contexte = contexte_test(ecole)
    if (config.provider or "").upper() == "META" and (config.template_meta or "").strip():
        message = resume_envoi_meta(config, contexte)
    else:
        nom_ecole = getattr(ecole, "ecole", None) or "Digital School"
        message = (
            f"Test Digital School — {nom_ecole}\n"
            f"Les notifications de paiement WhatsApp sont opérationnelles."
        )
        # Si texte libre Ultramsg, on peut aussi prévisualiser le modèle
        if (config.provider or "").upper() != "META":
            message = formater_message(config.modele_effectif(), contexte)

    ok, reponse, erreur = _envoyer_via_provider(
        config, telephone, message, contexte=contexte
    )
    NotificationWhatsApp.objects.create(
        ecole=ecole,
        paiement=None,
        destinataire=f"+{telephone}",
        message=message,
        statut="ENVOYE" if ok else "ECHEC",
        provider=config.provider,
        reponse_api=reponse or "",
        erreur=erreur or "",
    )
    if ok:
        messages.success(request, f"Message de test envoyé à +{telephone}.")
    else:
        messages.error(request, f"Échec d'envoi : {erreur or 'erreur API'}")
    return redirect("finances:whatsapp_config")


@login_required
def whatsapp_renvoyer(request, pk):
    """Renvoie la notification WhatsApp pour un paiement."""
    ecole = get_user_ecole(request)
    if not request.user.is_superuser:
        messages.error(request, "Seul un superutilisateur peut renvoyer un WhatsApp.")
        return redirect("finances:dashboard")

    paiement = get_object_or_404(paiements_for_ecole(ecole), pk=pk)
    if paiement.statut != "VALIDE":
        messages.error(request, "Seuls les paiements validés peuvent être notifiés.")
        return redirect("finances:whatsapp_config")

    from .whatsapp import notifier_paiement_whatsapp

    notif = notifier_paiement_whatsapp(paiement, force=True)
    if notif is None:
        messages.warning(
            request,
            "Aucune notification envoyée (WhatsApp inactif ou école introuvable).",
        )
    elif notif.statut == "ENVOYE":
        messages.success(request, f"WhatsApp renvoyé pour le reçu {paiement.numero_recu}.")
    elif notif.statut == "IGNORE":
        messages.warning(request, notif.erreur or "Notification ignorée.")
    else:
        messages.error(request, notif.erreur or "Échec d'envoi WhatsApp.")
    return redirect("finances:whatsapp_config")


@login_required
def paiement_list(request):
    ecole = get_user_ecole(request)
    qs = (
        paiements_for_ecole(ecole)
        .select_related(
            "eleve",
            "frais",
            "frais__type_frais",
            "devise",
            "devise_origine",
            "eleve__eleve",
        )
        .prefetch_related("demandes_modification")
        .order_by("-date_encodage")
    )

    # Caissier : uniquement ses propres encaissements
    filtre_perso = False
    if getattr(request.user, "is_caissier", False) and not request.user.is_superuser:
        qs = qs.filter(caissier__in=_identifiants_caissier(request.user))
        filtre_perso = True

    statut = request.GET.get("statut")
    if statut:
        qs = qs.filter(statut=statut)

    demandes_qs = DemandeModificationPaiement.objects.filter(
        paiement__eleve__eleve__ecole=ecole,
        statut="EN_ATTENTE",
    )
    if filtre_perso:
        demandes_qs = demandes_qs.filter(
            paiement__caissier__in=_identifiants_caissier(request.user)
        )
    demandes_ouvertes_ids = set(demandes_qs.values_list("paiement_id", flat=True))
    nb_demandes_ouvertes = len(demandes_ouvertes_ids)

    return render(
        request,
        "finances/paiement_list.html",
        {
            "object_list": qs,
            "demandes_ouvertes_ids": demandes_ouvertes_ids,
            "nb_demandes_ouvertes": nb_demandes_ouvertes,
            "filtre_perso": filtre_perso,
        },
    )


def _montant_recu_devise(paiement):
    """Montant réellement encaissé + code devise (origine si conversion)."""
    if paiement.montant_origine is not None and paiement.taux_change:
        code = (
            paiement.devise_origine.devise
            if paiement.devise_origine_id
            else "CDF"
        )
        return _decimal(paiement.montant_origine), code
    code = paiement.devise.devise if paiement.devise_id else "USD"
    return _decimal(paiement.montant_paye or 0), code


def _stats_paiements_journee(paiements):
    """Agrège les KPI d'une liste de paiements validés."""
    zero = _decimal("0")
    stats = {
        "nb_paiements": 0,
        "nb_eleves": 0,
        "total_usd": zero,
        "total_cdf": zero,
        "total_especes_usd": zero,
        "total_especes_cdf": zero,
        "total_mobile_usd": zero,
        "total_mobile_cdf": zero,
        "total_autres_usd": zero,
        "total_autres_cdf": zero,
    }
    eleves = set()
    for p in paiements:
        stats["nb_paiements"] += 1
        eleves.add(p.eleve_id)
        montant, code = _montant_recu_devise(p)
        mode = (p.mode_paiement or "").upper()
        if code == "USD":
            stats["total_usd"] += montant
            if mode == "ESPECES":
                stats["total_especes_usd"] += montant
            elif mode == "MOBILE_MONEY":
                stats["total_mobile_usd"] += montant
            else:
                stats["total_autres_usd"] += montant
        else:
            stats["total_cdf"] += montant
            if mode == "ESPECES":
                stats["total_especes_cdf"] += montant
            elif mode == "MOBILE_MONEY":
                stats["total_mobile_cdf"] += montant
            else:
                stats["total_autres_cdf"] += montant
    stats["nb_eleves"] = len(eleves)
    return stats


@caissier_required
def cloture_caisse(request):
    """Clôture de journée caissier + KPI une fois clôturée."""
    from datetime import datetime

    ecole = get_user_ecole(request)
    jour = timezone.localdate()
    date_param = (request.GET.get("date") or "").strip()
    if date_param:
        try:
            jour = datetime.strptime(date_param, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "Date invalide — journée du jour affichée.")

    noms = _identifiants_caissier(request.user)
    caissier_label = _nom_acteur(request.user)

    paiements = list(
        paiements_for_ecole(ecole)
        .filter(
            statut="VALIDE",
            caissier__in=noms,
            date_encodage__date=jour,
        )
        .select_related(
            "eleve",
            "eleve__eleve",
            "frais",
            "frais__type_frais",
            "devise",
            "devise_origine",
        )
        .order_by("date_encodage")
    )

    cloture = (
        ClotureCaisse.objects.filter(
            ecole=ecole,
            date_journee=jour,
            caissier__in=noms,
        )
        .order_by("-date_cloture")
        .first()
    )

    if request.method == "POST" and request.POST.get("action") == "cloturer":
        if cloture:
            messages.info(request, f"La journée du {jour:%d/%m/%Y} est déjà clôturée.")
            return redirect(
                f"{reverse('finances:cloture_caisse')}?date={jour.isoformat()}"
            )

        stats = _stats_paiements_journee(paiements)
        commentaire = (request.POST.get("commentaire") or "").strip()
        cloture = ClotureCaisse.objects.create(
            ecole=ecole,
            date_journee=jour,
            caissier=caissier_label,
            commentaire=commentaire,
            **stats,
        )
        messages.success(
            request,
            f"Journée du {jour:%d/%m/%Y} clôturée — {cloture.nb_paiements} paiement(s).",
        )
        return redirect(
            f"{reverse('finances:cloture_caisse')}?date={jour.isoformat()}&just_closed=1"
        )

    if cloture:
        kpis = {
            "nb_paiements": cloture.nb_paiements,
            "nb_eleves": cloture.nb_eleves,
            "total_usd": cloture.total_usd,
            "total_cdf": cloture.total_cdf,
            "total_especes_usd": cloture.total_especes_usd,
            "total_especes_cdf": cloture.total_especes_cdf,
            "total_mobile_usd": cloture.total_mobile_usd,
            "total_mobile_cdf": cloture.total_mobile_cdf,
            "total_autres_usd": cloture.total_autres_usd,
            "total_autres_cdf": cloture.total_autres_cdf,
        }
    else:
        kpis = _stats_paiements_journee(paiements)

    historique = (
        ClotureCaisse.objects.filter(ecole=ecole, caissier__in=noms)
        .order_by("-date_journee")[:14]
    )

    return render(
        request,
        "finances/cloture_caisse.html",
        {
            "jour": jour,
            "cloture": cloture,
            "kpis": kpis,
            "paiements": paiements,
            "historique": historique,
            "ecole": ecole,
            "just_closed": request.GET.get("just_closed") == "1" and cloture is not None,
        },
    )


@login_required
def budget_annuel(request):
    """Budget annuel complet : toutes rubriques + détail minerval par classe."""
    ecole = get_user_ecole(request)
    annees = annees_for_ecole(ecole).order_by("-est_encoure", "-anne_scolaire")
    annee_id = _parse_optional_int(request.GET.get("annee")) or _parse_optional_int(
        request.POST.get("annee")
    )
    annee = None
    if annee_id:
        annee = annees.filter(pk=annee_id).first()
    if annee is None:
        annee = annees.filter(est_encoure=True).first() or annees.first()

    budget = None
    if ecole and annee:
        budget = (
            BudgetAnnuel.objects.filter(ecole=ecole, annee=annee)
            .prefetch_related(
                "lignes__classe__section",
                "lignes__devise",
                "postes__rubrique",
            )
            .first()
        )

    # Affichage : montants figés si budget existant, sinon propositions auto
    plan = construire_postes_budget(ecole, annee, budget=None)
    projection = plan["projection_minerval"]

    if request.method == "POST" and request.POST.get("action") == "fixer":
        if not ecole or not annee:
            messages.error(request, "Impossible de fixer le budget : école ou année manquante.")
            return redirect(reverse("finances:budget_annuel"))

        with transaction.atomic():
            budget, _created = BudgetAnnuel.objects.get_or_create(
                ecole=ecole,
                annee=annee,
                defaults={
                    "date_fixation": timezone.now(),
                    "fixe_par": _nom_acteur(request.user),
                },
            )
            budget.date_fixation = timezone.now()
            budget.fixe_par = _nom_acteur(request.user)
            budget.commentaire = (request.POST.get("commentaire") or "").strip()
            budget.capacite_totale = projection["capacite_totale"]
            budget.total_usd = projection["total_usd"]
            budget.total_cdf = projection["total_cdf"]

            postes_payload = []
            total_r_usd = Decimal("0")
            total_r_cdf = Decimal("0")
            total_d_usd = Decimal("0")
            total_d_cdf = Decimal("0")

            for item in plan["postes"]:
                rub = item["rubrique"]
                usd = _decimal(request.POST.get(f"usd_{rub.id}", "0") or "0")
                cdf = _decimal(request.POST.get(f"cdf_{rub.id}", "0") or "0")
                note = (request.POST.get(f"note_{rub.id}") or "").strip()[:255]
                postes_payload.append(
                    PosteBudget(
                        budget=budget,
                        rubrique=rub,
                        montant_usd=usd,
                        montant_cdf=cdf,
                        est_auto=bool(rub.calcul_auto),
                        note=note or item["note"],
                    )
                )
                if rub.nature == "RECETTE":
                    total_r_usd += usd
                    total_r_cdf += cdf
                else:
                    total_d_usd += usd
                    total_d_cdf += cdf

            budget.total_recettes_usd = total_r_usd
            budget.total_recettes_cdf = total_r_cdf
            budget.total_depenses_usd = total_d_usd
            budget.total_depenses_cdf = total_d_cdf
            budget.save()

            budget.postes.all().delete()
            PosteBudget.objects.bulk_create(postes_payload)

            budget.lignes.all().delete()
            if projection["lignes"]:
                LigneBudget.objects.bulk_create(
                    [
                        LigneBudget(
                            budget=budget,
                            classe=ligne["classe"],
                            capacite=ligne["capacite"],
                            montant_unitaire=ligne["montant_unitaire"],
                            devise=ligne["devise"],
                            sous_total=ligne["sous_total"],
                            type_frais_libelle=ligne["type_frais_libelle"],
                        )
                        for ligne in projection["lignes"]
                    ]
                )

        messages.success(
            request,
            f"Budget {annee} fixé — recettes "
            f"{total_r_usd} USD / {total_r_cdf} CDF, dépenses "
            f"{total_d_usd} USD / {total_d_cdf} CDF.",
        )
        return redirect(f"{reverse('finances:budget_annuel')}?annee={annee.id}")

    # Recharger le plan avec montants figés pour l'affichage
    if budget:
        plan = construire_postes_budget(ecole, annee, budget=budget)

    realise = {"usd": Decimal("0"), "cdf": Decimal("0")}
    if ecole and annee:
        for row in (
            minerval_paiements_queryset(ecole, {"annee": annee.id})
            .values("devise__devise")
            .annotate(total=Sum("montant_paye"))
        ):
            code = (row["devise__devise"] or "").upper()
            if code == "USD":
                realise["usd"] = _decimal(row["total"])
            elif code == "CDF":
                realise["cdf"] = _decimal(row["total"])

    solde_usd = plan["total_recettes_usd"] - plan["total_depenses_usd"]
    solde_cdf = plan["total_recettes_cdf"] - plan["total_depenses_cdf"]
    taux_obj = TauxChange.courant_pour_ecole(ecole)

    return render(
        request,
        "finances/budget_annuel.html",
        {
            "ecole": ecole,
            "annees": annees,
            "annee": annee,
            "projection": projection,
            "plan": plan,
            "budget": budget,
            "realise": realise,
            "solde_usd": solde_usd,
            "solde_cdf": solde_cdf,
            "taux_courant": taux_obj,
        },
    )


@caissier_required
def demande_modification_create(request, pk):
    """Le caissier signale un paiement mal saisi pour correction par l'admin."""
    ecole = get_user_ecole(request)
    paiement = get_object_or_404(
        paiements_for_ecole(ecole).select_related(
            "eleve", "eleve__eleve", "frais", "frais__type_frais", "devise", "devise_origine"
        ),
        pk=pk,
    )
    existante = _demande_modification_ouverte(paiement)
    if existante:
        messages.info(
            request,
            f"Une demande de modification est déjà en attente pour le reçu {paiement.numero_recu}.",
        )
        return redirect("finances:demandes_modification_list")

    if request.method == "POST":
        form = DemandeModificationPaiementForm(request.POST)
        if form.is_valid():
            demande = form.save(commit=False)
            demande.paiement = paiement
            demande.demande_par = _nom_acteur(request.user)
            demande.statut = "EN_ATTENTE"
            demande.save()
            messages.success(
                request,
                f"Demande envoyée pour le reçu {paiement.numero_recu}. "
                "Un administrateur pourra corriger le paiement.",
            )
            return redirect("finances:demandes_modification_list")
    else:
        form = DemandeModificationPaiementForm()

    return render(
        request,
        "finances/demande_modification_form.html",
        {"form": form, "paiement": paiement},
    )


@login_required
def demandes_modification_list(request):
    """Liste des demandes — caissier : les siennes / admin : toutes celles de l'école."""
    ecole = get_user_ecole(request)
    qs = (
        DemandeModificationPaiement.objects.filter(paiement__eleve__eleve__ecole=ecole)
        .select_related(
            "paiement",
            "paiement__eleve",
            "paiement__eleve__eleve",
            "paiement__frais",
            "paiement__frais__type_frais",
            "paiement__devise",
            "paiement__devise_origine",
        )
        .order_by("-date_demande")
    )

    # Caissier : uniquement ses demandes (par nom d'acteur)
    if getattr(request.user, "is_caissier", False) and not request.user.is_superuser:
        qs = qs.filter(demande_par=_nom_acteur(request.user))

    statut = (request.GET.get("statut") or "EN_ATTENTE").strip()
    if statut in ("EN_ATTENTE", "TRAITEE", "REJETEE"):
        qs = qs.filter(statut=statut)
    elif statut != "TOUS":
        statut = "EN_ATTENTE"
        qs = qs.filter(statut="EN_ATTENTE")
    else:
        statut = "TOUS"

    return render(
        request,
        "finances/demandes_modification_list.html",
        {
            "object_list": qs,
            "filter_statut": statut,
            "nb_en_attente": DemandeModificationPaiement.objects.filter(
                paiement__eleve__eleve__ecole=ecole,
                statut="EN_ATTENTE",
            ).count(),
            "peut_traiter": _peut_modifier_paiement(request.user),
        },
    )


@modification_paiement_required
def demande_modification_rejeter(request, pk):
    ecole = get_user_ecole(request)
    demande = get_object_or_404(
        DemandeModificationPaiement.objects.filter(paiement__eleve__eleve__ecole=ecole),
        pk=pk,
        statut="EN_ATTENTE",
    )
    if request.method == "POST":
        demande.statut = "REJETEE"
        demande.traite_par = _nom_acteur(request.user)
        demande.date_traitement = timezone.now()
        demande.reponse_admin = (request.POST.get("reponse_admin") or "").strip()
        demande.save(
            update_fields=["statut", "traite_par", "date_traitement", "reponse_admin"]
        )
        messages.success(
            request,
            f"Demande rejetée pour le reçu {demande.paiement.numero_recu}.",
        )
        return redirect("finances:demandes_modification_list")
    return render(
        request,
        "finances/demande_modification_rejeter.html",
        {"demande": demande, "paiement": demande.paiement},
    )


@caissier_required
def paiement_create(request):
    ecole = get_user_ecole(request)
    taux_obj = TauxChange.courant_pour_ecole(ecole)
    taux_courant = taux_obj.taux if taux_obj else None
    if request.method == "POST":
        form = PaiementForm(request.POST, ecole=ecole, taux_courant=taux_courant)
        if form.is_valid():
            paiement = form.save(commit=False)

            if not paiement.numero_recu:
                year_month = timezone.now().strftime("%Y%m")
                last_paiement = (
                    paiements_for_ecole(ecole)
                    .filter(numero_recu__startswith=f"REC-{year_month}")
                    .order_by('id')
                    .last()
                )
                if last_paiement and last_paiement.numero_recu:
                    try:
                        last_seq = int(last_paiement.numero_recu.split("-")[-1])
                    except ValueError:
                        last_seq = 0
                else:
                    last_seq = 0
                paiement.numero_recu = f"REC-{year_month}-{last_seq + 1:04d}"

            paiement.save()

            if getattr(paiement, "statut", None) == "VALIDE":
                try:
                    _hoada_auto_post_payment_to_entries(paiement)
                except Exception:
                    pass
                messages.success(
                    request,
                    f"Paiement enregistré — reçu {paiement.numero_recu} pour {paiement.eleve.eleve.prenom} {paiement.eleve.eleve.nom}.",
                )
                # 2 exemplaires (école + parent), impression directe, retour à l'encaissement
                return redirect(
                    reverse("finances:paiement_print", kwargs={"pk": paiement.pk})
                    + "?autoprint=1&copies=2&next=create"
                )

            messages.success(request, "Paiement enregistré en attente de validation.")
            return redirect("finances:paiement_list")
    else:
        initial = {
            "caissier": getattr(request.user, "prenom", "") or request.user.get_full_name() or request.user.username,
            "statut": "VALIDE",
            "mode_paiement": "ESPECES",
        }
        inscription_id = request.GET.get("inscription")
        if inscription_id:
            ins = get_object_or_404(inscriptions_for_ecole(ecole), pk=inscription_id)
            initial["eleve"] = ins
        form = PaiementForm(
            initial=initial, ecole=ecole, exclude_paiement_id=None, taux_courant=taux_courant
        )

    context = _paiement_form_context(request, ecole, form, is_update=False)
    return render(request, "finances/paiement_form.html", context)


@login_required
def paiement_print(request, pk):
    ecole = get_user_ecole(request)
    paiement = get_object_or_404(
        paiements_for_ecole(ecole).select_related(
            "eleve", "eleve__eleve", "eleve__classe",
            "frais", "frais__type_frais", "frais__annee", "devise",
        ),
        pk=pk,
    )
    from .paiement_utils import paiements_valides_par_frais, solde_frais, _decimal

    paye_hors_ce = paiements_valides_par_frais(ecole, exclude_paiement_id=paiement.pk)
    solde_avant = solde_frais(paiement.frais, paiement.eleve_id, paye_hors_ce)
    montant_ce = _decimal(paiement.montant_paye)
    if paiement.statut == "VALIDE":
        reste_apres = max(solde_avant["reste"] - montant_ce, _decimal("0"))
        total_paye = solde_avant["paye"] + montant_ce
    else:
        reste_apres = solde_avant["reste"]
        total_paye = solde_avant["paye"]

    try:
        copies = int(request.GET.get("copies") or "1")
    except (TypeError, ValueError):
        copies = 1
    copies = 2 if copies >= 2 else 1
    exemplaires = (
        ["Exemplaire école", "Exemplaire parent"] if copies == 2 else [None]
    )
    next_param = (request.GET.get("next") or "").strip().lower()
    if next_param == "create":
        next_url = reverse("finances:paiement_create")
    else:
        next_url = reverse("finances:paiement_list")

    return render(
        request,
        "finances/paiement_print.html",
        {
            "paiement": paiement,
            "ecole": ecole,
            "autoprint": request.GET.get("autoprint") == "1",
            "copies": copies,
            "exemplaires": exemplaires,
            "next_url": next_url,
            "solde": {
                "total": solde_avant["total"],
                "deja_paye": solde_avant["paye"],
                "ce_paiement": montant_ce,
                "total_paye": total_paye,
                "reste_apres": reste_apres,
            },
        },
    )



# -------------------------
# HOADA helpers (seed + auto-post)
# -------------------------

def _hoada_get_or_seed_default_accounts_for_ecole(ecole):
    """
    Seed minimal HOADA accounts & journals for each ecole if missing.
    Mapping provided by the user:
    - créance: 411100
    - produit: 706000
    - trésorerie: depends on mode_paiement
      ESPECES -> 571100
      VIREMENT/CHEQUE -> 521100
      MOBILE_MONEY -> 565000
    """
    defaults = [
        {"numero": "411100", "libelle": "411 Élèves (Créance)", "devise": None},
        {"numero": "706000", "libelle": "706 Vente de services (Produit)", "devise": None},
        {"numero": "571100", "libelle": "571100 Caisse siège / caisse principale", "devise": None},
        {"numero": "521100", "libelle": "521100 Banques locales / devises nationales", "devise": None},
        {"numero": "565000", "libelle": "565000 Dépôts de monnaie électronique", "devise": None},
    ]

    for d in defaults:
        CompteComptable.objects.get_or_create(
            ecole=ecole,
            numero=d["numero"],
            defaults={"libelle": d["libelle"], "devise": d["devise"]},
        )


    journal_defaults = [
        {"code": "PAIE_ESPECES", "libelle": "Paiements - Espèces"},
        {"code": "PAIE_VIREMENT", "libelle": "Paiements - Virement"},
        {"code": "PAIE_CHEQUE", "libelle": "Paiements - Chèque"},
        {"code": "PAIE_MOBILE_MONEY", "libelle": "Paiements - Mobile Money"},
        {"code": "PAIE_SALAIRES", "libelle": "Paiements - Salaires"},
    ]
    for j in journal_defaults:
        JournalComptable.objects.get_or_create(
            ecole=ecole,
            code=j["code"],
            defaults={"libelle": j["libelle"]},
        )


def _hoada_get_account(ecole, numero):
    return CompteComptable.objects.filter(ecole=ecole, numero=str(numero)).first()


def _hoada_get_tresorerie_account(ecole, mode_paiement):
    mapping = {
        "ESPECES": "571100",
        "VIREMENT": "521100",
        "CHEQUE": "521100",
        "MOBILE_MONEY": "565000",
    }
    numero = mapping.get(mode_paiement)
    if not numero:
        return None
    return _hoada_get_account(ecole, numero)


def _hoada_get_journal(ecole, mode_paiement):
    code_map = {
        "ESPECES": "PAIE_ESPECES",
        "VIREMENT": "PAIE_VIREMENT",
        "CHEQUE": "PAIE_CHEQUE",
        "MOBILE_MONEY": "PAIE_MOBILE_MONEY",
    }
    code = code_map.get(mode_paiement)
    if not code:
        return None
    return JournalComptable.objects.filter(ecole=ecole, code=code).first()


def _hoada_payment_libelle(numero_recu):
    return f"Encaissement frais - Reçu {numero_recu}"


def _hoada_ecole_from_paiement(paiement):
    try:
        return paiement.eleve.classe.ecole
    except AttributeError:
        try:
            return getattr(paiement.eleve, "ecole", None) or getattr(
                getattr(paiement.frais, "section", None), "ecole", None
            )
        except AttributeError:
            return None


def _hoada_delete_payment_entries(ecole, numero_recu):
    """Supprime les écritures HOADA liées à un reçu (lignes + écriture)."""
    if not ecole or not numero_recu:
        return 0
    libelle = _hoada_payment_libelle(numero_recu)
    qs = Ecriture.objects.filter(ecole=ecole, libelle=libelle)
    count = qs.count()
    qs.delete()
    return count


def _hoada_auto_post_payment_to_entries(paiement, force=False):
    """
    Auto writing when payment becomes VALIDE:
    Encaissement des frais
    - Débit  : trésorerie (selon mode_paiement)
    - Crédit : 411100 (créance)

    Si force=True, recrée l'écriture même si elle existe déjà (après correction).
    Retourne True si une écriture est présente/créée, False sinon.
    """
    ecole = _hoada_ecole_from_paiement(paiement)
    if not ecole:
        return False

    numero = getattr(paiement, "numero_recu", None) or ""
    if not numero:
        return False

    if getattr(paiement, "statut", None) != "VALIDE":
        _hoada_delete_payment_entries(ecole, numero)
        return False

    _hoada_get_or_seed_default_accounts_for_ecole(ecole)

    compte_creance = _hoada_get_account(ecole, "411100")
    compte_treso = _hoada_get_tresorerie_account(ecole, paiement.mode_paiement)
    journal = _hoada_get_journal(ecole, paiement.mode_paiement)

    if not compte_creance or not compte_treso or not journal:
        return False

    montant = paiement.montant_paye if paiement.montant_paye is not None else getattr(paiement.frais, "montant", None)
    if not montant:
        return False

    libelle = _hoada_payment_libelle(numero)
    existing = Ecriture.objects.filter(ecole=ecole, libelle=libelle)
    if existing.exists():
        if not force:
            return True
        existing.delete()

    date_ecriture = (
        paiement.date_encodage.date()
        if getattr(paiement, "date_encodage", None)
        else timezone.now().date()
    )

    with transaction.atomic():
        ecriture = Ecriture.objects.create(
            ecole=ecole,
            date_ecriture=date_ecriture,
            journal=journal,
            piece=None,
            libelle=libelle,
        )
        EcritureLigne.objects.create(
            ecriture=ecriture,
            compte=compte_treso,
            sens="DEBIT",
            montant=montant,
        )
        EcritureLigne.objects.create(
            ecriture=ecriture,
            compte=compte_creance,
            sens="CREDIT",
            montant=montant,
        )
    return True


def _hoada_sync_payment_after_correction(paiement, ancien_numero_recu=None):
    """
    Après traitement d'une demande de modification :
    met à jour grand livre / balance / écritures pour ce paiement.
    """
    ecole = _hoada_ecole_from_paiement(paiement)
    if not ecole:
        return False

    if ancien_numero_recu and ancien_numero_recu != paiement.numero_recu:
        _hoada_delete_payment_entries(ecole, ancien_numero_recu)

    return bool(_hoada_auto_post_payment_to_entries(paiement, force=True))


def _hoada_auto_post_salary_to_entries(paie):
    """
    Auto writing when salary becomes PAYE:
    - Débit  : 660000 (Charges de personnel)
    - Crédit : trésorerie (selon mode_paiement)

    Retourne True si l'écriture existe (créée ou déjà présente), False sinon.
    """
    try:
        ecole = paie.personnel.ecole
    except AttributeError:
        return False

    if not ecole:
        return False

    _hoada_get_or_seed_default_accounts_for_ecole(ecole)

    compte_charges, _ = CompteComptable.objects.get_or_create(
        ecole=ecole,
        numero="660000",
        defaults={"libelle": "Charges de personnel (Salaires)", "devise": None},
    )
    compte_treso = _hoada_get_tresorerie_account(ecole, getattr(paie, "mode_paiement", "ESPECES"))
    journal, _ = JournalComptable.objects.get_or_create(
        ecole=ecole,
        code="PAIE_SALAIRES",
        defaults={"libelle": "Paiements - Salaires"},
    )

    if not compte_charges or not compte_treso or not journal:
        return False

    montant = paie.net_a_payer
    if not montant:
        return False

    ref_str = getattr(paie, "reference_paiement", None) or f"SAL-{paie.id}"
    libelle = f"Paiement Salaire {paie.mois}/{paie.annee} - {paie.personnel.nom} - Réf {ref_str}"

    if Ecriture.objects.filter(ecole=ecole, libelle=libelle).exists():
        return True

    date_ecriture = paie.date_paiement if getattr(paie, "date_paiement", None) else timezone.now().date()

    ecriture = Ecriture.objects.create(
        ecole=ecole,
        date_ecriture=date_ecriture,
        journal=journal,
        piece=None,
        libelle=libelle,
    )

    EcritureLigne.objects.create(
        ecriture=ecriture,
        compte=compte_charges,
        sens="DEBIT",
        montant=montant,
    )
    EcritureLigne.objects.create(
        ecriture=ecriture,
        compte=compte_treso,
        sens="CREDIT",
        montant=montant,
    )
    return True


@modification_paiement_required
def paiement_update(request, pk):
    ecole = get_user_ecole(request)
    obj = get_object_or_404(paiements_for_ecole(ecole), pk=pk)
    demande = _demande_modification_ouverte(obj)
    if not demande:
        messages.error(
            request,
            "Ce paiement ne peut être modifié que s'il existe une demande de modification "
            "en attente (déposée par le caissier).",
        )
        return redirect("finances:demandes_modification_list")

    taux_obj = TauxChange.courant_pour_ecole(ecole)
    taux_courant = taux_obj.taux if taux_obj else None
    if request.method == "POST":
        old_statut = obj.statut
        ancien_numero_recu = obj.numero_recu
        form = PaiementForm(
            request.POST,
            instance=obj,
            ecole=ecole,
            exclude_paiement_id=obj.pk,
            taux_courant=taux_courant,
        )
        if form.is_valid():
            # Seul un caissier peut valider / exécuter un paiement.
            if (
                old_statut != "VALIDE"
                and form.cleaned_data.get("statut") == "VALIDE"
                and not _peut_encaisser(request.user)
            ):
                messages.error(
                    request,
                    "Seul un caissier peut effectuer (valider) un paiement.",
                )
                context = _paiement_form_context(
                    request, ecole, form, is_update=True, paiement=obj
                )
                context["demande_modification"] = demande
                return render(request, "finances/paiement_update_form.html", context)

            updated = form.save()
            # Clôturer la demande après correction
            demande.statut = "TRAITEE"
            demande.traite_par = _nom_acteur(request.user)
            demande.date_traitement = timezone.now()
            if not demande.reponse_admin:
                demande.reponse_admin = "Paiement corrigé."
            demande.save(
                update_fields=["statut", "traite_par", "date_traitement", "reponse_admin"]
            )

            # Resynchroniser la comptabilité (écritures, grand livre, balance, caisse)
            ecriture_ok = False
            try:
                ecriture_ok = _hoada_sync_payment_after_correction(
                    updated, ancien_numero_recu=ancien_numero_recu
                )
            except Exception:
                ecriture_ok = False

            # Sync frais d'inscription (également via signal, on sécurise ici)
            try:
                from .paiement_utils import sync_frais_inscription_depuis_paiement
                sync_frais_inscription_depuis_paiement(updated)
            except Exception:
                pass

            if updated.statut == "VALIDE":
                msg = f"Paiement {updated.numero_recu} corrigé."
                if ecriture_ok:
                    msg += " Écriture comptable mise à jour (grand livre, balance, caisse)."
                else:
                    msg += " Attention : l'écriture comptable n'a pas pu être synchronisée."
                messages.success(request, msg)
                if old_statut != "VALIDE":
                    return redirect(
                        reverse("finances:paiement_print", kwargs={"pk": updated.pk})
                        + "?autoprint=1&copies=2&next=create"
                    )
            else:
                messages.success(
                    request,
                    f"Paiement {updated.numero_recu} corrigé (statut {updated.statut}). "
                    "Les écritures comptables liées ont été retirées si nécessaire.",
                )
            return redirect("finances:demandes_modification_list")
    else:
        form = PaiementForm(
            instance=obj,
            ecole=ecole,
            exclude_paiement_id=obj.pk,
            taux_courant=taux_courant,
        )
    context = _paiement_form_context(request, ecole, form, is_update=True, paiement=obj)
    context["demande_modification"] = demande
    return render(request, "finances/paiement_update_form.html", context)



@modification_paiement_required
def paiement_delete(request, pk):
    ecole = get_user_ecole(request)
    obj = get_object_or_404(paiements_for_ecole(ecole), pk=pk)
    if request.method == "POST":
        obj.delete()
        return redirect("finances:paiement_list")
    return render(request, "finances/paiement_confirm_delete.html", {"object": obj})


# -------------------------
# HOADA - Comptabilité
# -------------------------

@login_required
def hoada_plan_comptable_list(request):
    comptes = CompteComptable.objects.filter(ecole=request.user.ecole).order_by("numero")
    return render(request, "finances/hoada_plan_comptable_list.html", {"object_list": comptes})


@login_required
def hoada_plan_comptable_create(request):
    if request.method == "POST":
        form = CompteComptableForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = request.user.ecole
            obj.save()
            return redirect("finances:hoada_plan_comptable")
    else:
        form = CompteComptableForm()
    return render(request, "finances/hoada_plan_comptable_form.html", {"form": form})


@login_required
def hoada_journaux_list(request):
    journaux = JournalComptable.objects.filter(ecole=request.user.ecole).order_by("code")
    return render(request, "finances/hoada_journaux_list.html", {"object_list": journaux})


@login_required
def hoada_journaux_create(request):
    if request.method == "POST":
        form = JournalComptableForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = request.user.ecole
            obj.save()
            return redirect("finances:hoada_journaux_list")
    else:
        form = JournalComptableForm()
    return render(request, "finances/hoada_journaux_form.html", {"form": form})


@login_required
def hoada_ecritures_list(request):
    ecritures = (
        Ecriture.objects.filter(ecole=request.user.ecole)
        .select_related("journal", "piece")
        .order_by("-date_ecriture", "-id")
    )
    return render(request, "finances/hoada_ecritures_list.html", {"object_list": ecritures})


@login_required
def hoada_ecritures_create(request):
    if not _peut_gerer_ecritures(request.user):
        messages.error(
            request,
            "Accès réservé : seuls le trésorier et le comptable peuvent passer ou modifier les écritures.",
        )
        return redirect("finances:hoada_ecritures_list")

    # Saisie simplifiée : on crée l'en-tête + N lignes via request.POST (fields indexés)
    # Attendu côté template : compte_0, sens_0, montant_0 ... (jusqu'à line_count)
    if request.method == "POST":
        form_ecriture = EcritureForm(request.POST)
        try:
            line_count = int(request.POST.get("line_count", "0"))
        except ValueError:
            line_count = 0

        if form_ecriture.is_valid():
            ecriture = form_ecriture.save(commit=False)
            ecriture.ecole = request.user.ecole
            ecriture.save()

            ok = True
            lignes = []

            comptes_ids_autorises = set(
                CompteComptable.objects.filter(ecole=request.user.ecole).values_list("id", flat=True)
            )

            for i in range(line_count):
                compte_id = request.POST.get(f"compte_{i}")
                sens = request.POST.get(f"sens_{i}")
                montant = request.POST.get(f"montant_{i}")

                if not compte_id or not sens or not montant:
                    continue

                try:
                    compte_id_int = int(compte_id)
                except ValueError:
                    ok = False
                    break

                if compte_id_int not in comptes_ids_autorises:
                    ok = False
                    break

                ligne = EcritureLigne(
                    ecriture=ecriture,
                    compte_id=compte_id_int,
                    sens=sens,
                    montant=montant,
                )
                try:
                    ligne.full_clean()
                except Exception:
                    ok = False
                    break
                lignes.append(ligne)

            if not ok or not lignes:
                ecriture.delete()
                return render(
                    request,
                    "finances/hoada_ecritures_form.html",
                    {
                        "form": form_ecriture,
                        "line_count": line_count,
                        "comptes": CompteComptable.objects.filter(ecole=request.user.ecole).order_by("numero"),
                        "journaux": JournalComptable.objects.filter(ecole=request.user.ecole).order_by("code"),
                        "pieces": PieceComptable.objects.filter(ecole=request.user.ecole).order_by("-date", "reference"),
                        "error": "Vérifiez les lignes débit/crédit (compte, sens, montant) pour l'école courante.",
                    },
                )

            for l in lignes:
                l.save()

            return redirect("finances:hoada_ecritures_list")
    else:
        form_ecriture = EcritureForm()
        line_count = 3

    context = {
        "form": form_ecriture,
        "line_count": line_count,
        "comptes": CompteComptable.objects.filter(ecole=request.user.ecole).order_by("numero"),
        "journaux": JournalComptable.objects.filter(ecole=request.user.ecole).order_by("code"),
        "pieces": PieceComptable.objects.filter(ecole=request.user.ecole).order_by("-date", "reference"),
    }
    return render(request, "finances/hoada_ecritures_form.html", context)


@login_required
def hoada_grand_livre(request):
    lignes = (
        EcritureLigne.objects.select_related("ecriture", "compte", "ecriture__journal", "ecriture__piece")
        .filter(ecriture__ecole=request.user.ecole)
        .order_by("compte__numero", "ecriture__date_ecriture", "ecriture__id")
    )

    # Regroupe les lignes par compte avec totaux et solde
    comptes = []
    index = {}
    for l in lignes:
        numero = l.compte.numero
        if numero not in index:
            entry = {"compte": l.compte, "lignes": [], "debit": Decimal("0"), "credit": Decimal("0")}
            index[numero] = entry
            comptes.append(entry)
        entry = index[numero]
        entry["lignes"].append(l)
        if l.sens == "DEBIT":
            entry["debit"] += l.montant
        else:
            entry["credit"] += l.montant

    for entry in comptes:
        entry["solde"] = entry["debit"] - entry["credit"]

    total_debit = sum(e["debit"] for e in comptes) if comptes else Decimal("0")
    total_credit = sum(e["credit"] for e in comptes) if comptes else Decimal("0")

    return render(
        request,
        "finances/hoada_grand_livre.html",
        {
            "comptes": comptes,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "nb_lignes": lignes.count(),
        },
    )


@user_passes_test(lambda u: u.is_active and (u.is_staff or u.is_superuser))
def seed_finances(request):
    """Génère automatiquement TypeFrais/Frais_Scolaire/Paiement pour l'école courante."""
    from inscription.models import Inscription, Annee_Scolaire


    if request.method != "POST":
        return redirect("finances:dashboard")

    ecole = getattr(request.user, "ecole", None)
    if not ecole:
        return redirect("finances:dashboard")

    # Année scolaire: priorité à la currencue si trouvée, sinon première.
    annee_encours = Annee_Scolaire.objects.filter(est_encoure=True).first()
    if not annee_encours:
        annee_encours = Annee_Scolaire.objects.order_by("-anne_scolaire").first()
    if not annee_encours:
        return redirect("finances:dashboard")

    # Paramètres seed (valeurs simples et stables)
    default_devise_codes = ["USD", "CDF"]
    type_frais_defaults = [
        {"key": "minerval", "libelle": "Minerval", "description": "Frais scolaire"},
        {"key": "inscription", "libelle": "Frais inscription", "description": "Frais d'inscription"},
        {"key": "scolarite", "libelle": "Frais scolarité", "description": "Frais de scolarité"},
        {"key": "c", "libelle": "C", "description": "Frais C"},
    ]
    # Montants par défaut par type_frais
    montant_map = {
        "minerval": Decimal("500"),
        "inscription": Decimal("100"),
        "scolarite": Decimal("500"),
        "c": Decimal("200"),
    }

    from django.db import transaction

    with transaction.atomic():
        # Devise
        for code in default_devise_codes:
            from .models import Devise
            Devise.objects.get_or_create(devise=code)
        devise_usd = Devise.objects.filter(devise="USD").first()
        if not devise_usd:
            devise_usd = Devise.objects.first()

        # TypeFrais
        type_objs = {}
        for tf in type_frais_defaults:
            obj, _ = TypeFrais.objects.get_or_create(
                ecole=ecole,
                libelle=tf["libelle"],
                defaults={"description": tf["description"]},
            )
            type_objs[tf["key"]] = obj

        # Frais_Scolaire pour toutes les sections de l'école
        from inscription.tenant import sections_for_ecole
        sections = list(sections_for_ecole(ecole))
        # Modèle Annee_Scolaire: date_debut / date_fin
        echeance = getattr(annee_encours, "date_fin", None) or getattr(annee_encours, "date_debut")


        frais_by_type_section = {}
        for type_key, type_obj in type_objs.items():
            for section in sections:
                frais, _ = Frais_Scolaire.objects.get_or_create(
                    type_frais=type_obj,
                    annee=annee_encours,
                    section=section,
                    defaults={
                        "montant": montant_map.get(type_key, Decimal("0")),
                        "devise": devise_usd,
                        "echeance": echeance,
                        "est_obligatoire": True,
                    },
                )
                frais_by_type_section[(type_key, section.id)] = frais

        # Paiements VALIDE pour chaque inscription
        ins_qs = Inscription.objects.select_related("eleve", "classe").filter(
            annee_s=annee_encours, classe__ecole=ecole
        )
        # modes de paiement cycliques
        mode_cycle = ["ESPECES", "VIREMENT", "MOBILE_MONEY"]
        paiements_created = 0
        for idx, ins in enumerate(ins_qs):
            for type_key in ["inscription", "scolarite", "c"]:
                frais = frais_by_type_section.get((type_key, ins.classe.section_id))
                if not frais:
                    continue
                # numéro reçu déterministe
                numero_recu = f"{type_key[:3].upper()}-{ins.id}-{annee_encours.id}"
                if Paiement.objects.filter(numero_recu=numero_recu, frais=frais, eleve=ins).exists():
                    continue

                paiement = Paiement.objects.create(
                    eleve=ins,
                    frais=frais,
                    numero_recu=numero_recu,
                    montant_paye=frais.montant,
                    devise=frais.devise,
                    mode_paiement=mode_cycle[idx % len(mode_cycle)],
                    reference_trans=None,
                    caissier=str(getattr(request.user, "username", request.user.id)),
                    statut="VALIDE",
                )
                # Déclenche HOADA (logique existante)
                _hoada_auto_post_payment_to_entries(paiement)
                paiements_created += 1

    return redirect("finances:paiement_list")


@login_required
def inscriptions_non_paye(request):
    """Liste des inscriptions dont le frais d'inscription n'a pas de paiement VALIDE."""
    from inscription.models import Inscription

    ecole = get_user_ecole(request)

    types_inscription = type_frais_for_ecole(ecole).filter(libelle__icontains="inscription")
    if not types_inscription.exists():
        types_inscription = type_frais_for_ecole(ecole).filter(libelle__icontains="inscr")

    annee_encours = annees_for_ecole(ecole).filter(est_encoure=True).first()

    frais_qs = frais_for_ecole(ecole).select_related("type_frais", "annee", "section", "devise").filter(
        type_frais__in=types_inscription
    )
    if annee_encours:
        frais_qs = frais_qs.filter(annee=annee_encours)

    inscriptions_qs = (
        inscriptions_for_ecole(ecole)
        .select_related("eleve", "classe", "annee_s")
        .filter(annee_s__in=frais_qs.values_list("annee", flat=True).distinct())
    )

    paiement_valides = paiements_for_ecole(ecole).filter(statut="VALIDE")

    object_list = []
    for ins in inscriptions_qs:
        frais_ins = frais_qs.filter(annee=ins.annee_s, section_id=ins.classe.section_id)
        frais_ins = frais_ins.filter(type_frais__in=types_inscription)

        # Payé si au moins un paiement VALIDE existe
        ins.frais_inscription = paiement_valides.filter(eleve=ins, frais__in=frais_ins).exists()
        object_list.append(ins)


    return render(request, "finances/inscriptions_non_paye.html", {"object_list": object_list})


@login_required
def hoada_balance(request):
    lignes = (
        EcritureLigne.objects.select_related("compte")
        .filter(ecriture__ecole=request.user.ecole)
        .all()
    )

    balance = {}
    for l in lignes:
        numero = l.compte.numero
        if numero not in balance:
            balance[numero] = {"compte": l.compte, "debit": Decimal("0"), "credit": Decimal("0")}
        if l.sens == "DEBIT":
            balance[numero]["debit"] += l.montant
        else:
            balance[numero]["credit"] += l.montant

    rows = list(balance.values())
    rows.sort(key=lambda r: r["compte"].numero)

    for r in rows:
        solde = r["debit"] - r["credit"]
        r["solde_debiteur"] = solde if solde > 0 else Decimal("0")
        r["solde_crediteur"] = -solde if solde < 0 else Decimal("0")

    total_debit = sum(r["debit"] for r in rows) if rows else Decimal("0")
    total_credit = sum(r["credit"] for r in rows) if rows else Decimal("0")
    total_solde_debiteur = sum(r["solde_debiteur"] for r in rows) if rows else Decimal("0")
    total_solde_crediteur = sum(r["solde_crediteur"] for r in rows) if rows else Decimal("0")

    return render(
        request,
        "finances/hoada_balance.html",
        {
            "rows": rows,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "total_solde_debiteur": total_solde_debiteur,
            "total_solde_crediteur": total_solde_crediteur,
        },
    )


# -------------------------
# Salaires / Payer le personnel
# -------------------------

MOIS_LABELS = {
    1: "Janvier",
    2: "Février",
    3: "Mars",
    4: "Avril",
    5: "Mai",
    6: "Juin",
    7: "Juillet",
    8: "Août",
    9: "Septembre",
    10: "Octobre",
    11: "Novembre",
    12: "Décembre",
}


def marquer_paie_payee(paie, mode_paiement=None, date_paiement=None):
    """Marque une fiche de paie comme PAYÉE et poste l'écriture HOADA.

    Retourne un dict : {'payee': bool, 'ecriture_ok': bool, 'erreur': str|None}
    """
    result = {"payee": False, "ecriture_ok": False, "erreur": None}

    if paie.statut_paiement == "PAYE":
        # Déjà payée : vérifier / compléter l'écriture si manquante
        try:
            result["ecriture_ok"] = bool(_hoada_auto_post_salary_to_entries(paie))
        except Exception as exc:
            result["erreur"] = str(exc)
        return result

    if mode_paiement:
        modes_valides = {c[0] for c in paie.MODE_PAIEMENT_CHOICES}
        if mode_paiement in modes_valides:
            paie.mode_paiement = mode_paiement

    paie.statut_paiement = "PAYE"
    paie.date_paiement = date_paiement or date.today()
    if not paie.reference_paiement:
        paie.reference_paiement = f"SAL-{paie.annee}{paie.mois:02d}-{paie.id}"
    paie.save()
    result["payee"] = True

    # Le signal post_save tente déjà le post ; on sécurise ici aussi.
    try:
        result["ecriture_ok"] = bool(_hoada_auto_post_salary_to_entries(paie))
        if not result["ecriture_ok"]:
            result["erreur"] = (
                "Paiement enregistré, mais l'écriture comptable n'a pas pu être créée "
                "(compte de trésorerie ou journal manquant)."
            )
    except Exception as exc:
        result["erreur"] = f"Paiement enregistré, erreur comptable : {exc}"
    return result


@login_required
def salaires_list(request):
    """Liste des fiches de paie générées en GRH — paiement depuis Finances."""
    from django.db.models import Case, IntegerField, When

    from grh.models import Paie
    from grh.tenant import paies_for_ecole

    ecole = get_user_ecole(request)
    base_qs = paies_for_ecole(ecole).select_related(
        "personnel", "devise", "personnel__ecole"
    )

    statut = (request.GET.get("statut") or "TOUS").strip()
    qs = base_qs
    if statut in ("EN_ATTENTE", "PAYE", "ANNULE"):
        qs = qs.filter(statut_paiement=statut)
    else:
        statut = "TOUS"

    mois = _parse_optional_int(request.GET.get("mois"))
    annee = _parse_optional_int(request.GET.get("annee"))
    if mois:
        qs = qs.filter(mois=mois)
    if annee:
        qs = qs.filter(annee=annee)

    # En attente d'abord, puis les plus récentes
    qs = qs.annotate(
        statut_ordre=Case(
            When(statut_paiement="EN_ATTENTE", then=0),
            When(statut_paiement="PAYE", then=1),
            default=2,
            output_field=IntegerField(),
        )
    ).order_by(
        "statut_ordre",
        "-annee",
        "-mois",
        "personnel__nom",
        "personnel__Post_nom",
        "personnel__prenom",
        "-id",
    )

    total_en_attente = (
        base_qs.filter(statut_paiement="EN_ATTENTE")
        .values("devise__devise")
        .annotate(total=Sum("net_a_payer"))
        .order_by("-total")
    )

    annees_dispo = (
        base_qs.values_list("annee", flat=True).distinct().order_by("-annee")
    )

    return render(
        request,
        "finances/salaires_list.html",
        {
            "object_list": qs,
            "filter_statut": statut,
            "filter_mois": mois,
            "filter_annee": annee,
            "mois_labels": MOIS_LABELS,
            "modes_paiement": Paie.MODE_PAIEMENT_CHOICES,
            "total_en_attente": total_en_attente,
            "nb_en_attente": base_qs.filter(statut_paiement="EN_ATTENTE").count(),
            "nb_generees": base_qs.count(),
            "nb_payees": base_qs.filter(statut_paiement="PAYE").count(),
            "annees_dispo": list(annees_dispo),
            "ecole": ecole,
        },
    )


@login_required
def salaire_payer(request, pk):
    """Valide le paiement d'une fiche de paie depuis Finances."""
    from grh.models import Paie
    from grh.tenant import paies_for_ecole

    ecole = get_user_ecole(request)
    paie = get_object_or_404(paies_for_ecole(ecole), pk=pk)

    if request.method != "POST":
        return redirect("finances:salaires_list")

    if paie.statut_paiement == "PAYE":
        messages.info(request, "Cette fiche de paie est déjà payée.")
        return redirect("finances:salaires_list")

    if paie.statut_paiement == "ANNULE":
        messages.warning(request, "Impossible de payer une fiche annulée.")
        return redirect("finances:salaires_list")

    mode = (request.POST.get("mode_paiement") or "").strip() or None
    result = marquer_paie_payee(paie, mode_paiement=mode)
    if result["payee"]:
        if result["ecriture_ok"]:
            messages.success(
                request,
                f"Salaire de {paie.personnel.prenom} {paie.personnel.nom} "
                f"({paie.mois}/{paie.annee}) payé — {paie.net_a_payer} {paie.devise}. "
                f"Écriture comptable enregistrée (Débit 660000 / Crédit trésorerie).",
            )
        else:
            messages.warning(
                request,
                result["erreur"]
                or "Paiement enregistré, mais l'écriture comptable est manquante.",
            )
    elif result["erreur"]:
        messages.warning(request, result["erreur"])
    else:
        messages.info(request, "Cette fiche de paie est déjà payée.")
    return redirect("finances:salaires_list")


@login_required
def salaires_payer_lot(request):
    """Paiement groupé des fiches sélectionnées."""
    from grh.tenant import paies_for_ecole

    ecole = get_user_ecole(request)
    if request.method != "POST":
        return redirect("finances:salaires_list")

    ids = request.POST.getlist("paie_ids")
    mode = (request.POST.get("mode_paiement") or "").strip() or None
    qs = paies_for_ecole(ecole).filter(pk__in=ids, statut_paiement="EN_ATTENTE")

    payes = 0
    ecritures_ok = 0
    erreurs = []
    for paie in qs:
        result = marquer_paie_payee(paie, mode_paiement=mode)
        if result["payee"]:
            payes += 1
        if result["ecriture_ok"]:
            ecritures_ok += 1
        elif result["erreur"]:
            erreurs.append(result["erreur"])

    if payes:
        messages.success(
            request,
            f"{payes} fiche(s) payée(s), {ecritures_ok} écriture(s) comptable(s) enregistrée(s).",
        )
    else:
        messages.info(request, "Aucune fiche en attente sélectionnée.")
    for err in erreurs[:3]:
        messages.warning(request, err)
    return redirect("finances:salaires_list")

